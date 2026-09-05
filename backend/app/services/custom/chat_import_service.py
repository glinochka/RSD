"""Bulk import ChatTarget rows from CSV/XLSX files."""
from __future__ import annotations

import csv
import io
import logging
import re
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from ...alembic.models import ChatImportJob, ChatJoinStatus, ChatMode, ChatSource, ChatTarget
from ...config import settings
from .chat_target_dedup import dedup_keys, load_existing_dedup_keys
from .telegram_invite import TelegramChatRefError, parse_telegram_chat_ref

logger = logging.getLogger(__name__)

_PROVIDER = "telegram"
MAX_IMPORT_BYTES = 40 * 1024 * 1024
MAX_IMPORT_ROWS = 40_000
CHANNEL_MIN_SUBSCRIBERS = 100
CHAT_MIN_MEMBERS = 50
CHANNEL_MAX_IDLE = timedelta(days=60)
MAX_CELL_CHARS = 2048
MAX_TITLE_CHARS = 255
MAX_LINK_CHARS = 512
MAX_TYPE_CHARS = 32
MAX_DESCRIPTION_CHARS = 500
MAX_ERROR_LOG = 50
BATCH_COMMIT = 100
MAX_PLAUSIBLE_MEMBERS = 50_000_000

_LINK_KEYS = ("invite_link", "link", "url", "href", "ссылка")
_TITLE_KEYS = ("title", "name", "chat_name", "название", "имя")
_TYPE_KEYS = ("chat_type", "type", "тип")
_ID_KEYS = ("external_chat_id", "chat_id", "id")
_DESC_KEYS = ("description", "about", "описание")
_MEMBERS_KEYS = ("members", "members_count", "participants", "subscribers", "участники", "подписчики")
_ACTIVITY_KEYS = ("activity", "last_activity", "last_activity_at", "активность")
_TOPIC_KEYS = ("ключ", "key", "topic", "keyword")

_ACTIVITY_RE = re.compile(
    r"(\d+)\s*"
    r"(мин(?:ут[аы]?)?|час(?:а|ов)?|дн(?:я|ей)?|день|сут(?:ки|ок)?|недел[яиь]?|месяц(?:а|ев)?)",
    re.IGNORECASE,
)


def _utc_now() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


def _media_root() -> Path:
    return Path(settings.MEDIA_ROOT).resolve()


def _cell_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, float):
        if value != value:  # NaN
            return None
        if value.is_integer():
            value = int(value)
    text = str(value).strip()
    if not text or text.lower() in {"-", "nan", "nat", "null", "none", "n/a"}:
        return None
    if len(text) > MAX_CELL_CHARS:
        text = text[:MAX_CELL_CHARS]
    return text


def _normalize_header(raw: Any) -> str:
    text = str(raw or "").replace("\ufeff", "").strip().lower().replace("ё", "е")
    return " ".join(text.split())


def _row_get(row: dict[str, Any], keys: tuple[str, ...]) -> str | None:
    for key in keys:
        value = _cell_text(row.get(key))
        if value:
            return value
    for raw_key, raw_val in row.items():
        header = _normalize_header(raw_key)
        if any(header == key or header.startswith(f"{key} ") for key in keys):
            value = _cell_text(raw_val)
            if value:
                return value
    return None


def parse_members_count(raw: Any) -> int | None:
    """Drop Excel garbage (timestamps / huge ids). Keep plausible subscriber counts."""
    if raw is None or isinstance(raw, bool):
        return None
    if isinstance(raw, datetime):
        return None
    if isinstance(raw, float):
        if raw != raw:
            return None
        raw = int(raw) if raw.is_integer() else raw
    if isinstance(raw, int):
        n = raw
    else:
        text = _cell_text(raw)
        if not text:
            return None
        digits = re.sub(r"[^\d]", "", text)
        if not digits:
            return None
        try:
            n = int(digits)
        except ValueError:
            return None
    if n <= 0 or n > MAX_PLAUSIBLE_MEMBERS:
        return None
    return n


def parse_relative_activity(raw: Any, *, now: datetime | None = None) -> datetime | None:
    if isinstance(raw, datetime):
        value = raw
        if value.tzinfo is not None:
            value = value.astimezone(timezone.utc).replace(tzinfo=None)
        return value
    text = _cell_text(raw)
    if not text:
        return None
    match = _ACTIVITY_RE.search(text.replace(".", " "))
    if not match:
        return None
    amount = int(match.group(1))
    unit = match.group(2).lower()
    if unit.startswith("мин"):
        delta = timedelta(minutes=amount)
    elif unit.startswith("час"):
        delta = timedelta(hours=amount)
    elif unit.startswith("недел"):
        delta = timedelta(weeks=amount)
    elif unit.startswith("месяц"):
        delta = timedelta(days=30 * amount)
    else:
        delta = timedelta(days=amount)
    return (now or _utc_now()) - delta


def _normalize_invite(link: str | None) -> tuple[str | None, str | None]:
    raw = _cell_text(link)
    if not raw:
        return None, None
    try:
        parsed = parse_telegram_chat_ref(raw)
    except TelegramChatRefError:
        return raw[:MAX_LINK_CHARS], None
    external_id = parsed.value if parsed.kind == "channel_id" else None
    return parsed.canonical[:MAX_LINK_CHARS], external_id


def _dedup_key(invite_link: str | None, external_chat_id: str | None) -> str | None:
    if invite_link:
        return f"link:{invite_link.strip().lower()}"
    if external_chat_id:
        return f"id:{str(external_chat_id).strip().lower()}"
    return None


def _read_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for raw in reader:
        if len(rows) >= MAX_IMPORT_ROWS:
            break
        row = {_normalize_header(k): _cell_text(v) for k, v in raw.items() if k}
        rows.append(row)
    return rows


def _pick_excel_sheet(wb):
    for name in wb.sheetnames:
        if str(name).strip().lower() == "telegram":
            return wb[name]
    return wb.active


def _read_excel(content: bytes) -> list[dict[str, Any]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is not installed; cannot import Excel files") from exc

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = _pick_excel_sheet(wb)
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return []
        headers = [_normalize_header(cell) for cell in header_row]
        result: list[dict[str, Any]] = []
        for values in rows_iter:
            if len(result) >= MAX_IMPORT_ROWS:
                break
            row: dict[str, Any] = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                row[header] = values[idx] if idx < len(values) else None
            if any(_cell_text(v) for v in row.values()):
                result.append(row)
        return result
    finally:
        wb.close()


def _parse_rows(content: bytes, filename: str) -> list[dict[str, Any]]:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return _read_csv(content)
    if name.endswith((".xlsx", ".xls")):
        return _read_excel(content)
    try:
        return _read_csv(content)
    except Exception:
        return _read_excel(content)


def parse_import_row(row: dict[str, Any], *, now: datetime | None = None) -> dict[str, Any]:
    invite_link, invite_id = _normalize_invite(_row_get(row, _LINK_KEYS))
    external_chat_id = _cell_text(_row_get(row, _ID_KEYS)) or invite_id
    title = _cell_text(_row_get(row, _TITLE_KEYS))
    if title:
        title = title[:MAX_TITLE_CHARS]
    description = _cell_text(_row_get(row, _DESC_KEYS))
    topic = _cell_text(_row_get(row, _TOPIC_KEYS))
    if not description and topic:
        description = topic
    if description:
        description = description[:MAX_DESCRIPTION_CHARS]
    chat_type = _cell_text(_row_get(row, _TYPE_KEYS))
    if chat_type:
        chat_type = chat_type[:MAX_TYPE_CHARS].lower()
        if chat_type in {"канал", "channel", "broadcast"}:
            chat_type = "channel"
        elif chat_type in {"чат", "группа", "group", "megagroup", "chat"}:
            chat_type = "chat"
    members_count = parse_members_count(_row_get(row, _MEMBERS_KEYS) or row.get("участники"))
    if members_count is None:
        for key in _MEMBERS_KEYS:
            if key in row:
                members_count = parse_members_count(row.get(key))
                break
    last_activity_at = parse_relative_activity(
        _row_get(row, _ACTIVITY_KEYS) or row.get("активность"),
        now=now,
    )
    if not invite_link and not external_chat_id:
        raise ValueError("укажите ссылку или id чата/канала")
    keys = dedup_keys(invite_link, external_chat_id)
    return {
        "invite_link": invite_link,
        "external_chat_id": external_chat_id,
        "title": title,
        "description": description,
        "chat_type": chat_type,
        "members_count": members_count,
        "last_activity_at": last_activity_at,
        "dedup_key": next(iter(keys), None),
        "dedup_keys": keys,
    }


def import_quality_skip_reason(
    parsed: dict[str, Any],
    *,
    now: datetime | None = None,
) -> str | None:
    """Drop dead/tiny targets before they enter the join queue.

    Channels: last post older than 2 months, or fewer than 100 subscribers.
    Chats: fewer than 50 members.
    Missing Excel fields are not a fail — only known-bad rows are skipped.
    """
    current = now or _utc_now()
    chat_type = (parsed.get("chat_type") or "").strip().lower()
    members = parsed.get("members_count")
    activity = parsed.get("last_activity_at")
    stale = activity is not None and activity < (current - CHANNEL_MAX_IDLE)

    if chat_type == "channel":
        if members is not None and members < CHANNEL_MIN_SUBSCRIBERS:
            return f"канал: меньше {CHANNEL_MIN_SUBSCRIBERS} подписчиков ({members})"
        if stale:
            return "канал: последняя активность старше 2 месяцев"
        return None

    if chat_type == "chat":
        if members is not None and members < CHAT_MIN_MEMBERS:
            return f"чат: меньше {CHAT_MIN_MEMBERS} участников ({members})"
        return None

    if members is not None and members < CHAT_MIN_MEMBERS:
        return f"чат: меньше {CHAT_MIN_MEMBERS} участников ({members})"
    if stale and members is not None and members < CHANNEL_MIN_SUBSCRIBERS:
        return "канал: мало подписчиков и нет активности 2 месяца"
    return None


async def _existing_keys(session: AsyncSession, automation_id: int) -> set[str]:
    return await load_existing_dedup_keys(session, automation_id)


async def _save_import_file(automation_id: int, job_id: int, filename: str, content: bytes) -> str:
    root = _media_root()
    imports_dir = root / "chat_imports" / str(automation_id) / str(job_id)
    imports_dir.mkdir(parents=True, exist_ok=True)
    safe_name = f"{int(time.time())}_{Path(filename).name.replace(' ', '_')[:80]}"
    target = imports_dir / safe_name
    target.write_bytes(content)
    return str(target.relative_to(root))


def _append_error(errors: list[dict[str, Any]], row_idx: int, message: str) -> None:
    if len(errors) >= MAX_ERROR_LOG:
        return
    errors.append({"row": row_idx, "error": str(message)[:255]})


async def import_chats_from_file(
    session: AsyncSession,
    *,
    automation_id: int,
    filename: str,
    content: bytes,
    created_by_admin_id: int | None = None,
) -> ChatImportJob:
    if content is None:
        raise ValueError("empty file")
    if len(content) > MAX_IMPORT_BYTES:
        raise ValueError("Файл слишком большой (больше 40 МБ)")

    rows = _parse_rows(content, filename)
    job = ChatImportJob(
        custom_automation_id=automation_id,
        file_name=(filename or "import.xlsx")[:255],
        file_path=None,
        status="pending",
        total_rows=len(rows),
        processed_rows=0,
        error_rows=0,
        duplicate_rows=0,
        error_log=[],
        created_by_admin_id=created_by_admin_id,
        created_at=_utc_now(),
        updated_at=_utc_now(),
    )
    session.add(job)
    await session.flush()
    await session.refresh(job)

    file_path = await _save_import_file(automation_id, job.id, filename, content)
    job.file_path = file_path

    existing = await _existing_keys(session, automation_id)
    seen: set[str] = set()
    created = 0
    duplicates = 0
    skipped = 0
    errors: list[dict[str, Any]] = []
    filter_log: list[dict[str, Any]] = []
    now = _utc_now()

    for idx, row in enumerate(rows, start=1):
        try:
            parsed = parse_import_row(row, now=now)
            row_keys = set(parsed.get("dedup_keys") or [])
            if not row_keys:
                raise ValueError("укажите ссылку или id чата/канала")
            if row_keys & existing or row_keys & seen:
                duplicates += 1
                continue
            skip_reason = import_quality_skip_reason(parsed, now=now)
            if skip_reason:
                skipped += 1
                seen.update(row_keys)
                _append_error(filter_log, idx, f"фильтр: {skip_reason}")
                continue
            seen.update(row_keys)
            target = ChatTarget(
                custom_automation_id=automation_id,
                provider=_PROVIDER,
                external_chat_id=parsed["external_chat_id"],
                invite_link=parsed["invite_link"],
                title=parsed["title"],
                description=parsed["description"],
                chat_type=parsed["chat_type"],
                members_count=parsed["members_count"],
                last_activity_at=parsed["last_activity_at"],
                mode=ChatMode.MONITORING.value,
                source=ChatSource.BULK_IMPORT.value,
                import_job_id=job.id,
                join_status=ChatJoinStatus.PENDING.value,
                join_attempts=0,
                is_active=True,
                created_at=now,
                updated_at=now,
            )
            session.add(target)
            created += 1
            if created % BATCH_COMMIT == 0:
                await session.flush()
        except Exception as exc:
            _append_error(errors, idx, str(exc))

    job.processed_rows = created
    job.duplicate_rows = duplicates
    job.error_rows = len(errors)
    summary: list[dict[str, Any]] = []
    if skipped:
        summary.append(
            {
                "row": 0,
                "error": f"Отсеяно фильтром: {skipped}",
                "skipped_rows": skipped,
            }
        )
    job.error_log = summary + errors + filter_log
    if created == 0 and duplicates and not errors and not skipped:
        job.status = "completed"
    else:
        job.status = "completed" if not errors else "completed_with_errors"
    job.updated_at = _utc_now()
    await session.flush()
    from .chat_membership_service import ensure_memberships_for_automation

    await ensure_memberships_for_automation(session, automation_id)
    await session.commit()
    await session.refresh(job)
    return job


async def retry_import_errors(
    session: AsyncSession,
    job_id: int,
) -> ChatImportJob | None:
    job = await session.get(ChatImportJob, job_id)
    if not job or not job.file_path:
        return None
    content = (_media_root() / job.file_path).read_bytes()
    rows = _parse_rows(content, job.file_name)
    existing = await _existing_keys(session, job.custom_automation_id)
    seen: set[str] = set()
    created = 0
    now = _utc_now()
    remaining = []
    preserved = [
        item
        for item in (job.error_log or [])
        if isinstance(item, dict)
        and (
            item.get("skipped_rows") is not None
            or str(item.get("error") or "").startswith("фильтр:")
        )
    ]
    for idx, row in enumerate(rows, start=1):
        if not any(
            err.get("row") == idx
            and not str(err.get("error") or "").startswith("фильтр:")
            and err.get("skipped_rows") is None
            for err in (job.error_log or [])
        ):
            continue
        try:
            parsed = parse_import_row(row, now=now)
            row_keys = set(parsed.get("dedup_keys") or [])
            if not row_keys:
                continue
            if row_keys & existing or row_keys & seen:
                continue
            skip_reason = import_quality_skip_reason(parsed, now=now)
            if skip_reason:
                continue
            seen.update(row_keys)
            session.add(
                ChatTarget(
                    custom_automation_id=job.custom_automation_id,
                    provider=_PROVIDER,
                    external_chat_id=parsed["external_chat_id"],
                    invite_link=parsed["invite_link"],
                    title=parsed["title"],
                    description=parsed["description"],
                    chat_type=parsed["chat_type"],
                    members_count=parsed["members_count"],
                    last_activity_at=parsed["last_activity_at"],
                    mode=ChatMode.MONITORING.value,
                    source=ChatSource.BULK_IMPORT.value,
                    import_job_id=job.id,
                    join_status=ChatJoinStatus.PENDING.value,
                    join_attempts=0,
                    is_active=True,
                    created_at=now,
                    updated_at=now,
                )
            )
            created += 1
        except Exception as exc:
            logger.warning("Retry import row %s failed: %s", idx, exc)
            remaining.append({"row": idx, "error": str(exc)[:255]})

    job.processed_rows += created
    job.error_rows = len(remaining)
    job.error_log = preserved + remaining
    job.updated_at = _utc_now()
    from .chat_membership_service import ensure_memberships_for_automation

    await ensure_memberships_for_automation(session, job.custom_automation_id)
    await session.commit()
    await session.refresh(job)
    return job
