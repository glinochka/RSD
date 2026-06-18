"""Парсинг выгрузок контактов (2GIS и совместимые таблицы) в sales_outbound_contacts."""

from __future__ import annotations

import json
import re
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

_ws_re = re.compile(r"\s+")
PHONE_FIELD_MAX_LEN = 256
EMAIL_FIELD_MAX_LEN = 255
URL_FIELD_MAX_LEN = 512
NAME_FIELD_MAX_LEN = 256


def _norm_header(h: object) -> str:
    s = str(h or "").replace("\ufeff", "").strip()
    s = _ws_re.sub(" ", s)
    return s.casefold()


def _cell_str(val: object) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def _opt_str(val: object) -> str | None:
    s = _cell_str(val)
    return s or None


def _fit_phone(value: str | None, *, max_len: int = PHONE_FIELD_MAX_LEN) -> str | None:
    if not value:
        return None
    if len(value) <= max_len:
        return value
    return value[:max_len]


def _fit_str(value: str | None, *, max_len: int) -> str | None:
    if not value:
        return None
    if len(value) <= max_len:
        return value
    return value[:max_len]


def _fit_email(value: str | None, *, max_len: int = EMAIL_FIELD_MAX_LEN) -> str | None:
    """2GIS часто кладёт несколько адресов в одну ячейку — в БД varchar(255), берём первый."""
    s = (value or "").strip()
    if not s:
        return None
    for part in re.split(r"[,;]\s*", s):
        candidate = part.strip()
        if candidate and "@" in candidate:
            s = candidate
            break
    return _fit_str(s, max_len=max_len)


def _normalize_whatsapp_import_value(value: str | None) -> str | None:
    """Исправляет wa.me/17… (лишняя 1 перед RU +7) и приводит номер к https://wa.me/<digits>."""
    s = (value or "").strip()
    if not s:
        return None
    m = re.match(r"^https?://(?:www\.)?wa\.me/([^/?#]+)", s, re.I)
    phone_raw = m.group(1) if m else s
    digits = re.sub(r"\D", "", phone_raw)
    if len(digits) < 10:
        return s
    if len(digits) >= 12 and digits.startswith("17"):
        digits = digits[1:]
    if len(digits) == 11 and digits.startswith("8"):
        digits = "7" + digits[1:]
    elif len(digits) == 10 and digits.startswith("9"):
        digits = "7" + digits
    return f"https://wa.me/{digits}"


def _layout_yandex_maps(headers: list[str]) -> dict[str, int] | None:
    """Выгрузка Яндекс Карт: ID, Название, Регион, Город, Адрес, …, Рубрика, Подрубрика, …"""
    n = [_norm_header(h) for h in headers]
    if len(n) < 11 or n[0] != "id" or n[1] != "название":
        return None
    if n[10] != "рубрика":
        return None
    mapping: dict[str, int] = {
        "org_name": 1,
        "region": 2,
        "city": 3,
        "org_address": 4,
        "postal_code": 5,
        "org_phone": 6,
        "org_mobile": 7,
        "email": 8,
        "website": 9,
        "category_rubric": 10,
        "category_subrubric": 11,
    }
    for i, ni in enumerate(n):
        if ni == "время работы":
            mapping["working_hours"] = i
        elif ni == "способы оплаты":
            mapping["payment_methods"] = i
        elif ni == "whatsapp":
            mapping["whatsapp"] = i
        elif ni == "telegram":
            mapping["telegram"] = i
        elif ni == "рейтинг":
            mapping["rating"] = i
        elif ni == "кол-во отзывов":
            mapping["reviews_count"] = i
    return mapping


def _layout_by_position(headers: list[str]) -> dict[str, int] | None:
    """Стандартный порядок колонок из выгрузки (Название, ФИО ЛПР, Телефон, ...)."""
    n = [_norm_header(h) for h in headers]
    if len(n) < 6:
        return None
    ok0 = n[0] == "название" or n[0].startswith("название")
    ok1 = "фио" in n[1] and "лпр" in n[1]
    ok2 = n[2] == "телефон"
    ok3 = "мобильн" in n[3]
    ok4 = "телефон" in n[4] and "лпр" in n[4]
    if not (ok0 and ok1 and ok2 and ok3 and ok4):
        return None
    return {
        "org_name": 0,
        "lpr_name": 1,
        "org_phone": 2,
        "org_mobile": 3,
        "lpr_phone": 4,
        "import_status": 5,
    }


def _extend_mapping(headers: list[str], base: dict[str, int]) -> dict[str, int]:
    n = [_norm_header(h) for h in headers]
    used = set(base.values())
    m = dict(base)
    for i, ni in enumerate(n):
        if i in used:
            continue
        if ni in ("сайт", "website"):
            m.setdefault("website", i)
        elif ni in ("email", "e-mail", "e_mail", "почта"):
            m.setdefault("email", i)
        elif ni == "рубрика":
            m.setdefault("category_rubric", i)
        elif ni == "подрубрика":
            m.setdefault("category_subrubric", i)
        elif ni == "регион":
            m.setdefault("region", i)
        elif ni == "город":
            m.setdefault("city", i)
        elif ni == "адрес":
            m.setdefault("org_address", i)
        elif ni == "whatsapp" or "whatsapp" in ni:
            m.setdefault("whatsapp", i)
        elif ni == "telegram" or "телеграм" in ni:
            m.setdefault("telegram", i)
        elif ni == "макс" or (ni.startswith("макс") and "максим" not in ni):
            m.setdefault("messenger_max", i)
    return m


def _fuzzy_mapping(headers: list[str]) -> dict[str, int]:
    n = [_norm_header(h) for h in headers]
    m: dict[str, int] = {}
    for i, (raw, ni) in enumerate(zip(headers, n)):
        if "фио" in ni and "лпр" in ni:
            m.setdefault("lpr_name", i)
        elif ni == "название" and "(eng)" not in str(raw).lower() and "англ" not in ni:
            m.setdefault("org_name", i)
        elif "мобильн" in ni or ni == "мобильный телефон":
            m.setdefault("org_mobile", i)
        elif "телефон" in ni and "лпр" in ni:
            m.setdefault("lpr_phone", i)
        elif ni == "телефон" or ni.startswith("телефон "):
            if "лпр" not in ni:
                m.setdefault("org_phone", i)
        elif ni == "статус":
            m.setdefault("import_status", i)
        elif ni in ("сайт", "website"):
            m.setdefault("website", i)
        elif ni in ("email", "e-mail", "e_mail", "почта"):
            m.setdefault("email", i)
        elif ni == "рубрика":
            m.setdefault("category_rubric", i)
        elif ni == "подрубрика":
            m.setdefault("category_subrubric", i)
        elif ni == "регион":
            m.setdefault("region", i)
        elif ni == "город":
            m.setdefault("city", i)
        elif ni == "адрес":
            m.setdefault("org_address", i)
        elif ni == "whatsapp" or "whatsapp" in ni:
            m.setdefault("whatsapp", i)
        elif ni == "telegram" or "телеграм" in ni:
            m.setdefault("telegram", i)
        elif ni == "макс" or (ni.startswith("макс") and "максим" not in ni):
            m.setdefault("messenger_max", i)
    return _extend_mapping(headers, m)


def _pick_mapping(headers: list[str]) -> dict[str, int]:
    yandex = _layout_yandex_maps(headers)
    if yandex is not None:
        return _extend_mapping(headers, yandex)
    by_pos = _layout_by_position(headers)
    if by_pos is not None:
        return _extend_mapping(headers, by_pos)
    return _fuzzy_mapping(headers)


def parse_sales_excel(file_bytes: bytes) -> list[dict[str, Any]]:
    bio = BytesIO(file_bytes)
    wb = load_workbook(bio, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return []
        headers = [_cell_str(c) for c in header_row]
        colmap = _pick_mapping(headers)
        out: list[dict[str, Any]] = []
        for row in rows:
            raw_vals = list(row)
            while len(raw_vals) < len(headers):
                raw_vals.append(None)
            vals = raw_vals[: len(headers)]
            extras: dict[str, str] = {}
            picked: dict[str, str] = {}
            for key, idx in colmap.items():
                if idx < len(vals):
                    picked[key] = _cell_str(vals[idx])
            used_indices = set(colmap.values())
            for j, h in enumerate(headers):
                if j >= len(vals):
                    continue
                if j in used_indices:
                    continue
                hh = _cell_str(h)
                if hh:
                    extras[hh] = _cell_str(vals[j])
            org_name = picked.get("org_name", "").strip()
            if not org_name:
                col0 = _cell_str(vals[0]) if vals else ""
                if col0:
                    org_name = col0
                else:
                    continue
            email_raw = _opt_str(picked.get("email"))
            email = _fit_email(email_raw)
            if email_raw and email and email_raw.strip() != email:
                extras.setdefault("Email (полный из выгрузки)", email_raw)
            region = _fit_str(_opt_str(picked.get("region")), max_len=256)
            city = _fit_str(_opt_str(picked.get("city")), max_len=256)
            org_address = _fit_str(_opt_str(picked.get("org_address")), max_len=512)
            rubric = _fit_str(_opt_str(picked.get("category_rubric")), max_len=512)
            subrubric = _fit_str(_opt_str(picked.get("category_subrubric")), max_len=512)
            if region:
                extras.setdefault("region", region)
            if city:
                extras.setdefault("city", city)
            if org_address:
                extras.setdefault("address", org_address)
            if rubric:
                extras.setdefault("rubric", rubric)
                extras.setdefault("рубрика", rubric)
            if subrubric:
                extras.setdefault("subrubric", subrubric)
                extras.setdefault("подрубрика", subrubric)
            for extra_key, pick_key in (
                ("working_hours", "working_hours"),
                ("payment_methods", "payment_methods"),
                ("rating", "rating"),
                ("reviews_count", "reviews_count"),
            ):
                val = _opt_str(picked.get(pick_key))
                if val:
                    extras.setdefault(extra_key, val)
            out.append(
                {
                    "org_name": org_name[:512],
                    "lpr_name": _fit_str(_opt_str(picked.get("lpr_name")), max_len=NAME_FIELD_MAX_LEN),
                    "lpr_phone": _fit_phone(_opt_str(picked.get("lpr_phone"))),
                    "org_phone": _fit_phone(_opt_str(picked.get("org_phone"))),
                    "org_mobile": _fit_phone(_opt_str(picked.get("org_mobile"))),
                    "import_status": _opt_str(picked.get("import_status")),
                    "email": email,
                    "region": region,
                    "city": city,
                    "org_address": org_address,
                    "category_rubric": rubric,
                    "category_subrubric": subrubric,
                    "website": _fit_str(_opt_str(picked.get("website")), max_len=URL_FIELD_MAX_LEN),
                    "whatsapp": _fit_str(
                        _normalize_whatsapp_import_value(_opt_str(picked.get("whatsapp"))),
                        max_len=URL_FIELD_MAX_LEN,
                    ),
                    "telegram": _fit_str(_opt_str(picked.get("telegram")), max_len=URL_FIELD_MAX_LEN),
                    "messenger_max": _fit_str(_opt_str(picked.get("messenger_max")), max_len=URL_FIELD_MAX_LEN),
                    "extras": extras,
                }
            )
        return out
    finally:
        wb.close()


def extras_to_json(extras: dict[str, str]) -> str | None:
    if not extras:
        return None
    return json.dumps({"import_columns": extras}, ensure_ascii=False)
